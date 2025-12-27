"""
Tests for analyze_timelapse.py

Tests the analysis script's ability to:
- Find and match images with metadata files
- Load and parse metadata
- Calculate image brightness
- Extract EXIF data
- Find transition zones
- Generate graphs
- Export to Excel
- CLI main function
"""

import pytest
import json
import tempfile
import shutil
import os
from pathlib import Path
from datetime import datetime, timedelta
from unittest.mock import patch, MagicMock
import yaml
from PIL import Image
import numpy as np

# Import functions from analyze_timelapse
import sys

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
from analyze_timelapse import (
    load_config,
    find_recent_images,
    load_metadata,
    analyze_images,
    print_statistics,
    export_to_excel,
    calculate_image_brightness,
    extract_exif_data,
    find_transition_zones,
    create_graphs,
    main,
)


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test files."""
    temp_path = Path(tempfile.mkdtemp())
    yield temp_path
    shutil.rmtree(temp_path)


@pytest.fixture
def sample_config(temp_dir):
    """Create a sample config file."""
    config = {
        "output": {"directory": str(temp_dir / "images")},
        "graphs": {
            "directory": str(temp_dir / "graphs"),
            "width": 14,
            "height": 8,
            "dpi": 150,
            "default_hours": 24,
        },
        "adaptive_timelapse": {
            "light_thresholds": {"night": 10, "day": 100},
            "night_mode": {"max_exposure_time": 20.0},
        },
    }

    config_path = temp_dir / "config.yml"
    with open(config_path, "w") as f:
        yaml.dump(config, f)

    return config_path


@pytest.fixture
def sample_images_with_metadata(temp_dir):
    """Create sample images and metadata files."""
    images_dir = temp_dir / "images" / "2025" / "11" / "07"
    images_dir.mkdir(parents=True, exist_ok=True)

    image_metadata_pairs = []
    base_time = datetime.now()

    # Create 10 sample images with metadata
    for i in range(10):
        timestamp = base_time - timedelta(minutes=i * 5)

        # Create image file (empty, we don't actually analyze it)
        img_name = f"test_{timestamp.strftime('%Y_%m_%d_%H_%M_%S')}.jpg"
        img_path = images_dir / img_name
        img_path.write_bytes(b"\xff\xd8\xff\xe0")  # Minimal JPEG header

        # Create metadata file
        metadata = {
            "capture_timestamp": timestamp.isoformat(),
            "Lux": 1000 + (i * 100),
            "ExposureTime": 5000 + (i * 1000),  # microseconds
            "AnalogueGain": 1.5 + (i * 0.1),
            "SensorTemperature": 45.0 + i,
            "ColourTemperature": 6500 + (i * 100),
            "ColourGains": [1.5, 1.3],
            "DigitalGain": 1.0,
        }

        meta_name = f"test_{timestamp.strftime('%Y_%m_%d_%H_%M_%S')}_metadata.json"
        meta_path = images_dir / meta_name
        with open(meta_path, "w") as f:
            json.dump(metadata, f)

        # Set modification times to match the timestamp for proper sorting
        mtime = timestamp.timestamp()
        os.utime(img_path, (mtime, mtime))
        os.utime(meta_path, (mtime, mtime))

        image_metadata_pairs.append((img_path, meta_path))

    return images_dir, image_metadata_pairs


class TestLoadConfig:
    """Tests for load_config function."""

    def test_load_valid_config(self, sample_config):
        """Test loading a valid config file."""
        config = load_config(sample_config)

        assert "output" in config
        assert "graphs" in config
        assert config["graphs"]["width"] == 14
        assert config["adaptive_timelapse"]["light_thresholds"]["night"] == 10

    def test_load_nonexistent_config(self, temp_dir):
        """Test loading a non-existent config file."""
        with pytest.raises(FileNotFoundError):
            load_config(temp_dir / "nonexistent.yml")


class TestFindRecentImages:
    """Tests for find_recent_images function."""

    def test_find_all_images(self, sample_images_with_metadata):
        """Test finding all images within time window."""
        images_dir, expected_pairs = sample_images_with_metadata

        # Find images from last 24 hours
        found_pairs = find_recent_images(images_dir.parent.parent.parent, hours=24)

        # Should find all 10 test images
        assert len(found_pairs) == 10

        # Check that pairs are sorted chronologically
        timestamps = [datetime.fromtimestamp(p[0].stat().st_mtime) for p in found_pairs]
        assert timestamps == sorted(timestamps), "Images should be sorted chronologically"

    def test_find_limited_timeframe(self, sample_images_with_metadata):
        """Test finding images within a limited time window."""
        images_dir, _ = sample_images_with_metadata

        # Find images from last 15 minutes (should get ~3 images)
        found_pairs = find_recent_images(images_dir.parent.parent.parent, hours=0.25)

        # Should find fewer images (or equal if all created within window)
        assert len(found_pairs) <= 10
        assert len(found_pairs) > 0

    def test_empty_directory(self, temp_dir):
        """Test finding images in an empty directory."""
        empty_dir = temp_dir / "empty"
        empty_dir.mkdir()

        found_pairs = find_recent_images(empty_dir, hours=24)

        assert len(found_pairs) == 0


class TestLoadMetadata:
    """Tests for load_metadata function."""

    def test_load_valid_metadata(self, sample_images_with_metadata):
        """Test loading valid metadata file."""
        _, pairs = sample_images_with_metadata
        _, meta_path = pairs[0]

        metadata = load_metadata(meta_path)

        assert "Lux" in metadata
        assert "ExposureTime" in metadata
        assert isinstance(metadata["Lux"], (int, float))

    def test_load_nonexistent_metadata(self, temp_dir):
        """Test loading non-existent metadata file."""
        fake_path = temp_dir / "nonexistent_metadata.json"

        metadata = load_metadata(fake_path)

        # Should return empty dict on error
        assert metadata == {}

    def test_load_invalid_json(self, temp_dir):
        """Test loading invalid JSON file."""
        bad_json = temp_dir / "bad_metadata.json"
        bad_json.write_text("{ invalid json }")

        metadata = load_metadata(bad_json)

        # Should return empty dict on error
        assert metadata == {}


class TestAnalyzeImages:
    """Tests for analyze_images function."""

    def test_analyze_valid_images(self, sample_images_with_metadata):
        """Test analyzing valid images with metadata."""
        _, pairs = sample_images_with_metadata

        data = analyze_images(pairs, hours=24)

        # Check that all expected keys are present
        assert "timestamps" in data
        assert "lux" in data
        assert "exposure_time" in data
        assert "analogue_gain" in data
        assert "sensor_temp" in data

        # Check data length
        assert len(data["timestamps"]) == 10
        assert len(data["lux"]) == 10

        # Check that lux values are in expected range
        assert all(1000 <= lux <= 2000 for lux in data["lux"])

        # Check that exposure times are converted to seconds
        assert all(0.001 < exp < 1.0 for exp in data["exposure_time"])

    def test_analyze_empty_list(self):
        """Test analyzing empty list of images."""
        data = analyze_images([], hours=24)

        assert len(data["timestamps"]) == 0
        assert len(data["lux"]) == 0


class TestPrintStatistics:
    """Tests for print_statistics function."""

    def test_print_valid_statistics(self, capsys):
        """Test printing statistics with valid data."""
        data = {
            "timestamps": [datetime.now() - timedelta(hours=i) for i in range(5)],
            "lux": [100, 500, 1000, 5000, 10000],
            "exposure_time": [0.001, 0.01, 0.1, 1.0, 2.0],
            "analogue_gain": [1.0, 1.5, 2.0, 2.5, 3.0],
            "sensor_temp": [40, 45, 50, 55, 60],
            "colour_temp": [5000, 6000, 7000, 8000, 9000],
        }

        print_statistics(data, hours=24)

        captured = capsys.readouterr()
        assert "STATISTICAL SUMMARY" in captured.out
        assert "Light Levels" in captured.out
        assert "Exposure Time" in captured.out

    def test_print_empty_statistics(self, capsys):
        """Test printing statistics with empty data."""
        data = {
            "timestamps": [],
            "lux": [],
            "exposure_time": [],
            "analogue_gain": [],
            "sensor_temp": [],
            "colour_temp": [],
        }

        print_statistics(data, hours=24)

        # Should handle empty data gracefully (no output)
        captured = capsys.readouterr()
        assert captured.out == ""


class TestExportToExcel:
    """Tests for export_to_excel function."""

    def test_export_valid_data(self, temp_dir, sample_config, sample_images_with_metadata):
        """Test exporting valid data to Excel."""
        _, pairs = sample_images_with_metadata

        data = {
            "timestamps": [datetime.now() - timedelta(hours=i) for i in range(5)],
            "lux": [100, 500, 1000, 5000, 10000],
            "exposure_time": [0.001, 0.01, 0.1, 1.0, 2.0],
            "analogue_gain": [1.0, 1.5, 2.0, 2.5, 3.0],
            "sensor_temp": [40, 45, 50, 55, 60],
            "colour_temp": [5000, 6000, 7000, 8000, 9000],
            "colour_gains_red": [1.5, 1.6, 1.7, 1.8, 1.9],
            "colour_gains_blue": [1.3, 1.4, 1.5, 1.6, 1.7],
            "digital_gain": [1.0, 1.0, 1.0, 1.0, 1.0],
            "filenames": [f"test_{i}.jpg" for i in range(5)],
            "mode": ["day", "day", "day", "day", "day"],
            # Diagnostic fields (can be None if diagnostics disabled)
            "raw_lux": [None] * 5,
            "smoothed_lux": [None] * 5,
            "target_exposure_ms": [None] * 5,
            "interpolated_exposure_ms": [None] * 5,
            "target_gain": [None] * 5,
            "interpolated_gain": [None] * 5,
            "transition_position": [None] * 5,
            "brightness_mean": [None] * 5,
            "brightness_median": [None] * 5,
            "brightness_std": [None] * 5,
            "brightness_p5": [None] * 5,
            "brightness_p95": [None] * 5,
            "underexposed_percent": [None] * 5,
            "overexposed_percent": [None] * 5,
        }

        config = load_config(sample_config)
        output_path = temp_dir / "test_export.xlsx"

        export_to_excel(data, output_path, hours=24, config=config, image_pairs=pairs[:5])

        # Check that file was created
        assert output_path.exists()
        assert output_path.stat().st_size > 0

        # Try to open with openpyxl to verify it's valid
        from openpyxl import load_workbook

        wb = load_workbook(output_path)

        # Check that all sheets exist
        assert "Raw Data" in wb.sheetnames
        assert "Statistics" in wb.sheetnames
        assert "Hourly Averages" in wb.sheetnames

        # Check Raw Data sheet has data
        ws = wb["Raw Data"]
        assert ws.max_row >= 6  # Header + 5 data rows

        wb.close()


class TestEndToEnd:
    """End-to-end integration tests."""

    def test_complete_analysis_workflow(self, temp_dir, sample_config, sample_images_with_metadata):
        """Test the complete analysis workflow."""
        images_dir, pairs = sample_images_with_metadata
        config = load_config(sample_config)

        # 1. Find images
        found_pairs = find_recent_images(images_dir.parent.parent.parent, hours=24)
        assert len(found_pairs) == 10

        # 2. Analyze images
        data = analyze_images(found_pairs, hours=24)
        assert len(data["timestamps"]) == 10

        # 3. Export to Excel
        output_path = temp_dir / "complete_test.xlsx"
        export_to_excel(data, output_path, hours=24, config=config, image_pairs=found_pairs)
        assert output_path.exists()

        # 4. Verify Excel contents
        from openpyxl import load_workbook

        wb = load_workbook(output_path)

        ws_raw = wb["Raw Data"]
        assert ws_raw.max_row == 11  # Header + 10 data rows

        # Verify data is chronologically sorted
        timestamps = [ws_raw.cell(row=i, column=1).value for i in range(2, 12)]
        assert timestamps == sorted(timestamps), "Excel data should be chronologically sorted"

        wb.close()


class TestCalculateImageBrightness:
    """Tests for calculate_image_brightness function."""

    def test_calculate_brightness_white_image(self, temp_dir):
        """Test brightness calculation for white image."""
        img_path = temp_dir / "white.jpg"
        img = Image.new("RGB", (100, 100), color=(255, 255, 255))
        img.save(img_path, "JPEG")

        brightness = calculate_image_brightness(img_path)

        assert brightness > 250  # Should be close to 255

    def test_calculate_brightness_black_image(self, temp_dir):
        """Test brightness calculation for black image."""
        img_path = temp_dir / "black.jpg"
        img = Image.new("RGB", (100, 100), color=(0, 0, 0))
        img.save(img_path, "JPEG")

        brightness = calculate_image_brightness(img_path)

        assert brightness < 5  # Should be close to 0

    def test_calculate_brightness_gray_image(self, temp_dir):
        """Test brightness calculation for gray image."""
        img_path = temp_dir / "gray.jpg"
        img = Image.new("RGB", (100, 100), color=(128, 128, 128))
        img.save(img_path, "JPEG")

        brightness = calculate_image_brightness(img_path)

        assert 120 < brightness < 135  # Should be close to 128

    def test_calculate_brightness_invalid_file(self, temp_dir):
        """Test brightness calculation for invalid image."""
        img_path = temp_dir / "invalid.jpg"
        img_path.write_bytes(b"not an image")

        brightness = calculate_image_brightness(img_path)

        assert brightness == 0.0  # Should return 0 on error

    def test_calculate_brightness_nonexistent_file(self, temp_dir):
        """Test brightness calculation for non-existent file."""
        img_path = temp_dir / "nonexistent.jpg"

        brightness = calculate_image_brightness(img_path)

        assert brightness == 0.0  # Should return 0 on error


class TestExtractExifData:
    """Tests for extract_exif_data function."""

    def test_extract_exif_no_data(self, temp_dir):
        """Test EXIF extraction from image without EXIF."""
        img_path = temp_dir / "no_exif.jpg"
        img = Image.new("RGB", (100, 100), color=(100, 100, 100))
        img.save(img_path, "JPEG")

        exif = extract_exif_data(img_path)

        assert exif == {}

    def test_extract_exif_invalid_file(self, temp_dir):
        """Test EXIF extraction from invalid file."""
        img_path = temp_dir / "invalid.jpg"
        img_path.write_bytes(b"not an image")

        exif = extract_exif_data(img_path)

        assert exif == {}

    def test_extract_exif_nonexistent_file(self, temp_dir):
        """Test EXIF extraction from non-existent file."""
        img_path = temp_dir / "nonexistent.jpg"

        exif = extract_exif_data(img_path)

        assert exif == {}


class TestFindTransitionZones:
    """Tests for find_transition_zones function."""

    def test_find_zones_single_mode(self):
        """Test finding zones with single mode."""
        timestamps = [datetime.now() - timedelta(hours=i) for i in range(5)]
        modes = ["day", "day", "day", "day", "day"]

        zones = find_transition_zones(timestamps, modes)

        assert len(zones) == 1
        assert zones[0][2] == "day"

    def test_find_zones_two_modes(self):
        """Test finding zones with mode change."""
        base_time = datetime.now()
        timestamps = [base_time - timedelta(hours=i) for i in range(6)]
        modes = ["day", "day", "day", "night", "night", "night"]

        zones = find_transition_zones(timestamps, modes)

        assert len(zones) == 2
        assert zones[0][2] == "day"
        assert zones[1][2] == "night"

    def test_find_zones_multiple_transitions(self):
        """Test finding zones with multiple transitions."""
        base_time = datetime.now()
        timestamps = [base_time - timedelta(hours=i) for i in range(8)]
        modes = ["night", "night", "transition", "day", "day", "transition", "night", "night"]

        zones = find_transition_zones(timestamps, modes)

        assert len(zones) == 5

    def test_find_zones_empty(self):
        """Test finding zones with empty lists."""
        zones = find_transition_zones([], [])

        assert zones == []

    def test_find_zones_none_handling(self):
        """Test finding zones handles None modes."""
        base_time = datetime.now()
        timestamps = [base_time - timedelta(hours=i) for i in range(3)]
        modes = [None, None, None]

        zones = find_transition_zones(timestamps, modes)

        assert len(zones) == 1
        assert zones[0][2] == "unknown"


class TestCreateGraphs:
    """Tests for create_graphs function."""

    def test_create_graphs_empty_data(self, temp_dir, capsys):
        """Test create_graphs with empty data."""
        data = {"timestamps": [], "lux": []}
        config = {"adaptive_timelapse": {"light_thresholds": {"night": 10, "day": 100}}}

        create_graphs(data, temp_dir / "graphs", config)

        captured = capsys.readouterr()
        assert "No data to plot" in captured.out

    def test_create_graphs_with_data(self, temp_dir, sample_config):
        """Test create_graphs creates output files."""
        config = load_config(sample_config)

        # Create sample data
        base_time = datetime.now()
        data = {
            "timestamps": [base_time - timedelta(hours=i) for i in range(24)],
            "lux": [1000 + i * 100 for i in range(24)],
            "exposure_time": [0.01 + i * 0.001 for i in range(24)],
            "analogue_gain": [1.0 + i * 0.1 for i in range(24)],
            "sensor_temp": [45.0 + i for i in range(24)],
            "colour_temp": [5000 + i * 100 for i in range(24)],
            "colour_gains_red": [1.5] * 24,
            "colour_gains_blue": [1.3] * 24,
            "digital_gain": [1.0] * 24,
            "mode": ["day"] * 12 + ["transition"] * 4 + ["night"] * 8,
            "raw_lux": [None] * 24,
            "smoothed_lux": [None] * 24,
            "target_exposure_ms": [None] * 24,
            "interpolated_exposure_ms": [None] * 24,
            "target_gain": [None] * 24,
            "interpolated_gain": [None] * 24,
            "transition_position": [None] * 24,
            "sun_elevation": [None] * 24,
            "brightness_mean": [None] * 24,
            "brightness_median": [None] * 24,
            "brightness_std": [None] * 24,
            "brightness_p5": [None] * 24,
            "brightness_p95": [None] * 24,
            "underexposed_percent": [None] * 24,
            "overexposed_percent": [None] * 24,
        }

        output_dir = temp_dir / "graphs"
        create_graphs(data, output_dir, config)

        # Check that graphs were created
        assert output_dir.exists()
        graph_files = list(output_dir.glob("*.png"))
        assert len(graph_files) > 0


class TestMainCLI:
    """Tests for main CLI function."""

    def test_main_help(self, monkeypatch, capsys):
        """Test main with --help."""
        monkeypatch.setattr("sys.argv", ["analyze_timelapse.py", "--help"])

        with pytest.raises(SystemExit) as exc_info:
            main()

        assert exc_info.value.code == 0

    def test_main_nonexistent_config(self, temp_dir, monkeypatch, capsys):
        """Test main with non-existent config."""
        monkeypatch.setattr(
            "sys.argv",
            ["analyze_timelapse.py", "-c", str(temp_dir / "nonexistent.yml")],
        )

        with pytest.raises(SystemExit) as exc_info:
            main()

        assert exc_info.value.code == 1

    def test_main_valid_config_no_images(self, sample_config, temp_dir, monkeypatch, capsys):
        """Test main with valid config but no images."""
        # Create empty images directory
        images_dir = temp_dir / "images"
        images_dir.mkdir(parents=True, exist_ok=True)

        monkeypatch.setattr(
            "sys.argv",
            ["analyze_timelapse.py", "-c", str(sample_config), "--hours", "24"],
        )

        with pytest.raises(SystemExit) as exc_info:
            main()

        captured = capsys.readouterr()
        assert "No images found" in captured.out
        assert exc_info.value.code == 1

    def test_main_custom_hours(self, sample_config, monkeypatch):
        """Test main with custom hours parameter."""
        monkeypatch.setattr(
            "sys.argv",
            ["analyze_timelapse.py", "-c", str(sample_config), "--hours", "48"],
        )

        # Should exit with code 1 since there are no images
        with pytest.raises(SystemExit) as exc_info:
            main()
        assert exc_info.value.code == 1


class TestAnalyzeImagesWithDiagnostics:
    """Tests for analyze_images with diagnostic data."""

    def test_analyze_images_with_diagnostics(self, temp_dir):
        """Test analyzing images that have diagnostic metadata."""
        images_dir = temp_dir / "images"
        images_dir.mkdir(parents=True)

        base_time = datetime.now()

        # Create image and metadata with diagnostics
        img_path = images_dir / "test_image.jpg"
        img = Image.new("RGB", (100, 100), color=(100, 100, 100))
        img.save(img_path, "JPEG")

        metadata = {
            "capture_timestamp": base_time.isoformat(),
            "Lux": 1500,
            "ExposureTime": 10000,
            "AnalogueGain": 2.0,
            "SensorTemperature": 45.0,
            "ColourTemperature": 6500,
            "ColourGains": [1.5, 1.3],
            "DigitalGain": 1.0,
            "diagnostics": {
                "mode": "day",
                "raw_lux": 1500.0,
                "smoothed_lux": 1480.0,
                "target_exposure_ms": 10.0,
                "interpolated_exposure_ms": 10.5,
                "target_gain": 1.5,
                "interpolated_gain": 1.6,
                "transition_position": 0.5,
                "sun_elevation": 15.0,
                "brightness": {
                    "mean_brightness": 120.5,
                    "median_brightness": 118.0,
                    "std_brightness": 25.0,
                    "percentile_5": 45.0,
                    "percentile_95": 195.0,
                    "underexposed_percent": 2.5,
                    "overexposed_percent": 1.0,
                },
            },
        }

        meta_path = images_dir / "test_image_metadata.json"
        with open(meta_path, "w") as f:
            json.dump(metadata, f)

        # Set file times
        mtime = base_time.timestamp()
        os.utime(img_path, (mtime, mtime))
        os.utime(meta_path, (mtime, mtime))

        pairs = [(img_path, meta_path)]
        data = analyze_images(pairs, hours=24)

        # Check diagnostic data was captured
        assert len(data["mode"]) == 1
        assert data["mode"][0] == "day"
        assert data["raw_lux"][0] == 1500.0
        assert data["brightness_mean"][0] == 120.5

    def test_analyze_images_missing_timestamp(self, temp_dir):
        """Test analyzing images with missing capture_timestamp."""
        images_dir = temp_dir / "images"
        images_dir.mkdir(parents=True)

        base_time = datetime.now()

        # Create image
        img_path = images_dir / "test_image.jpg"
        img = Image.new("RGB", (100, 100), color=(100, 100, 100))
        img.save(img_path, "JPEG")

        # Create metadata WITHOUT capture_timestamp
        metadata = {
            "Lux": 1000,
            "ExposureTime": 5000,
            "AnalogueGain": 1.5,
        }

        meta_path = images_dir / "test_image_metadata.json"
        with open(meta_path, "w") as f:
            json.dump(metadata, f)

        # Set file times
        mtime = base_time.timestamp()
        os.utime(img_path, (mtime, mtime))
        os.utime(meta_path, (mtime, mtime))

        pairs = [(img_path, meta_path)]
        data = analyze_images(pairs, hours=24)

        # Should use file modification time as fallback
        assert len(data["timestamps"]) == 1


class TestExportToExcelWithDiagnostics:
    """Tests for export_to_excel with diagnostic data."""

    def test_export_with_full_diagnostics(self, temp_dir, sample_config):
        """Test Excel export with complete diagnostic data."""
        config = load_config(sample_config)

        base_time = datetime.now()
        data = {
            "timestamps": [base_time - timedelta(hours=i) for i in range(3)],
            "lux": [1000, 500, 100],
            "exposure_time": [0.01, 0.05, 0.1],
            "analogue_gain": [1.0, 2.0, 4.0],
            "sensor_temp": [45, 46, 47],
            "colour_temp": [6500, 6000, 5500],
            "colour_gains_red": [1.5, 1.6, 1.7],
            "colour_gains_blue": [1.3, 1.4, 1.5],
            "digital_gain": [1.0, 1.0, 1.0],
            "filenames": ["img1.jpg", "img2.jpg", "img3.jpg"],
            "mode": ["day", "transition", "night"],
            "raw_lux": [1000.0, 500.0, 100.0],
            "smoothed_lux": [980.0, 490.0, 95.0],
            "target_exposure_ms": [10.0, 50.0, 100.0],
            "interpolated_exposure_ms": [10.5, 55.0, 100.0],
            "target_gain": [1.0, 2.0, 4.0],
            "interpolated_gain": [1.0, 2.2, 4.0],
            "transition_position": [None, 0.5, None],
            "brightness_mean": [120.0, 100.0, 80.0],
            "brightness_median": [118.0, 98.0, 78.0],
            "brightness_std": [20.0, 25.0, 30.0],
            "brightness_p5": [40.0, 30.0, 20.0],
            "brightness_p95": [200.0, 180.0, 160.0],
            "underexposed_percent": [1.0, 2.0, 5.0],
            "overexposed_percent": [0.5, 0.3, 0.1],
        }

        # Create dummy image pairs
        images_dir = temp_dir / "images"
        images_dir.mkdir(parents=True)
        pairs = []
        for i in range(3):
            img_path = images_dir / f"img{i}.jpg"
            meta_path = images_dir / f"img{i}_metadata.json"
            img_path.touch()
            meta_path.touch()
            pairs.append((img_path, meta_path))

        output_path = temp_dir / "diagnostics_export.xlsx"
        export_to_excel(data, output_path, hours=24, config=config, image_pairs=pairs)

        assert output_path.exists()

        from openpyxl import load_workbook

        wb = load_workbook(output_path)
        assert "Raw Data" in wb.sheetnames
        ws = wb["Raw Data"]
        assert ws.max_row >= 4  # Header + 3 data rows
        wb.close()


class TestFindRecentImagesMetadataFolder:
    """Tests for find_recent_images with metadata in special folder."""

    def test_excludes_metadata_folder(self, temp_dir):
        """Test that images in 'metadata' folder are excluded."""
        images_dir = temp_dir / "images"
        metadata_folder = images_dir / "metadata"
        metadata_folder.mkdir(parents=True)

        base_time = datetime.now()

        # Create regular image
        img_path = images_dir / "regular.jpg"
        meta_path = images_dir / "regular_metadata.json"
        img_path.write_bytes(b"\xff\xd8\xff\xe0")

        metadata = {"capture_timestamp": base_time.isoformat(), "Lux": 1000}
        with open(meta_path, "w") as f:
            json.dump(metadata, f)

        mtime = base_time.timestamp()
        os.utime(img_path, (mtime, mtime))
        os.utime(meta_path, (mtime, mtime))

        # Create test shot in metadata folder
        test_img = metadata_folder / "test_shot.jpg"
        test_meta = metadata_folder / "test_shot_metadata.json"
        test_img.write_bytes(b"\xff\xd8\xff\xe0")
        with open(test_meta, "w") as f:
            json.dump(metadata, f)

        os.utime(test_img, (mtime, mtime))
        os.utime(test_meta, (mtime, mtime))

        # Find images - should only find regular, not test shot
        found_pairs = find_recent_images(images_dir, hours=24)

        assert len(found_pairs) == 1
        assert found_pairs[0][0].name == "regular.jpg"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

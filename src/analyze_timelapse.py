#!/usr/bin/env python3
"""
Raspilapse Analysis Script

Analyzes timelapse images and metadata from the last 24 hours.
Generates graphs showing:
- Lux levels over time
- Image brightness over time
- Exposure time over time
- Analogue gain (ISO) over time
- Sensor temperature over time
- Color temperature and gains over time

Usage:
    python3 src/analyze_timelapse.py [-c CONFIG] [--hours HOURS]

Examples:
    # Analyze last 24 hours (default)
    python3 src/analyze_timelapse.py

    # Analyze last 48 hours
    python3 src/analyze_timelapse.py --hours 48

    # Use custom config
    python3 src/analyze_timelapse.py -c config/custom.yml
"""

import argparse
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib

matplotlib.use("Agg")  # Non-interactive backend for headless systems
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import yaml
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def load_config(config_path: str) -> dict:
    """Load configuration from YAML file."""
    with open(config_path, "r") as f:
        return yaml.safe_load(f)


def find_recent_images(output_dir: Path, hours: int = 24) -> List[Tuple[Path, Path]]:
    """
    Find all images captured in the last N hours and match with metadata files.

    Metadata files may have different timestamps than images due to processing delay,
    so we match by finding the nearest metadata file within 60 seconds.

    Returns list of tuples: (image_path, metadata_path) sorted chronologically.
    """
    cutoff_time = datetime.now() - timedelta(hours=hours)

    # First, collect all JSON metadata files with their timestamps
    json_files = {}  # {mtime: path}
    for json_path in output_dir.rglob("*_metadata.json"):
        # Skip test shots in metadata folder
        # Skip only if parent dir name is exactly "metadata"
        if json_path.parent.name != "metadata":
            mtime = datetime.fromtimestamp(json_path.stat().st_mtime)
            if mtime >= cutoff_time:
                json_files[mtime] = json_path

    # Sort JSON file times for efficient lookup
    sorted_json_times = sorted(json_files.keys())

    # Now find JPG files and match with nearest metadata
    image_metadata_pairs = []

    for img_path in output_dir.rglob("*.jpg"):
        # Skip test shots in metadata folder
        if img_path.parent.name == "metadata":
            continue

        # Check if file was modified within the time window
        img_mtime = datetime.fromtimestamp(img_path.stat().st_mtime)
        if img_mtime < cutoff_time:
            continue

        # Find nearest metadata file (within 60 seconds)
        best_match = None
        best_diff = timedelta(seconds=60)

        for json_mtime in sorted_json_times:
            time_diff = abs(json_mtime - img_mtime)
            if time_diff < best_diff:
                best_diff = time_diff
                best_match = json_files[json_mtime]
            elif best_match and time_diff > best_diff:
                # We found a match and times are getting worse, so stop
                break

        if best_match:
            image_metadata_pairs.append((img_path, best_match))

    # Sort by image file modification time (chronologically from earliest to latest)
    image_metadata_pairs.sort(key=lambda x: x[0].stat().st_mtime)

    return image_metadata_pairs


def calculate_image_brightness(image_path: Path) -> float:
    """
    Calculate average brightness of an image.

    Returns brightness value 0-255 (average of all pixels in grayscale).
    """
    try:
        with Image.open(image_path) as img:
            # Convert to grayscale
            grayscale = img.convert("L")
            # Calculate average pixel value
            pixels = np.array(grayscale)
            return float(np.mean(pixels))
    except Exception as e:
        print(f"Warning: Could not calculate brightness for {image_path}: {e}")
        return 0.0


def extract_exif_data(image_path: Path) -> Dict:
    """
    Extract EXIF data from image.

    Returns dict of EXIF tags.
    """
    try:
        with Image.open(image_path) as img:
            exif = img.getexif()
            if exif:
                return {k: v for k, v in exif.items()}
            return {}
    except Exception as e:
        print(f"Warning: Could not extract EXIF from {image_path}: {e}")
        return {}


def load_metadata(metadata_path: Path) -> Dict:
    """Load metadata from JSON file."""
    try:
        with open(metadata_path, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"Warning: Could not load metadata from {metadata_path}: {e}")
        return {}


def analyze_images(image_metadata_pairs: List[Tuple[Path, Path]], hours: int) -> Dict:
    """
    Analyze metadata from all images and collect data.

    Returns dict with lists of data points for plotting.
    """
    data = {
        "timestamps": [],
        "lux": [],
        "exposure_time": [],  # in seconds
        "analogue_gain": [],
        "sensor_temp": [],
        "colour_temp": [],
        "colour_gains_red": [],
        "colour_gains_blue": [],
        "digital_gain": [],
        "filenames": [],
        # Diagnostic data
        "mode": [],
        "raw_lux": [],
        "smoothed_lux": [],
        "target_exposure_ms": [],
        "interpolated_exposure_ms": [],
        "target_gain": [],
        "interpolated_gain": [],
        "transition_position": [],
        "sun_elevation": [],  # Sun position in degrees (polar awareness)
        # Brightness analysis
        "brightness_mean": [],
        "brightness_median": [],
        "brightness_std": [],
        "brightness_p5": [],
        "brightness_p95": [],
        "underexposed_percent": [],
        "overexposed_percent": [],
    }

    num_images = len(image_metadata_pairs)
    print(f"\n📊 Analyzing metadata from {num_images} images from last {hours} hours...")

    for i, (img_path, meta_path) in enumerate(image_metadata_pairs, 1):
        if i % 100 == 0:
            print(f"  Processed {i}/{len(image_metadata_pairs)} metadata files...")

        # Load metadata
        metadata = load_metadata(meta_path)
        if not metadata:
            continue

        # Parse timestamp
        try:
            timestamp_str = metadata.get("capture_timestamp", "")
            timestamp = datetime.fromisoformat(timestamp_str)
        except:
            # Fallback to file modification time
            timestamp = datetime.fromtimestamp(img_path.stat().st_mtime)

        # Collect metadata
        data["timestamps"].append(timestamp)
        data["lux"].append(metadata.get("Lux", 0))
        data["filenames"].append(img_path.name)

        # Convert exposure time from microseconds to seconds
        exposure_us = metadata.get("ExposureTime", 0)
        data["exposure_time"].append(exposure_us / 1_000_000)

        data["analogue_gain"].append(metadata.get("AnalogueGain", 0))
        data["sensor_temp"].append(metadata.get("SensorTemperature", 0))
        data["colour_temp"].append(metadata.get("ColourTemperature", 0))

        # Color gains (red, blue)
        colour_gains = metadata.get("ColourGains", [0, 0])
        data["colour_gains_red"].append(colour_gains[0] if len(colour_gains) > 0 else 0)
        data["colour_gains_blue"].append(colour_gains[1] if len(colour_gains) > 1 else 0)

        data["digital_gain"].append(metadata.get("DigitalGain", 1.0))

        # Collect diagnostic data if available
        diagnostics = metadata.get("diagnostics", {})
        data["mode"].append(diagnostics.get("mode", "unknown"))
        data["raw_lux"].append(diagnostics.get("raw_lux"))
        data["smoothed_lux"].append(diagnostics.get("smoothed_lux"))
        data["target_exposure_ms"].append(diagnostics.get("target_exposure_ms"))
        data["interpolated_exposure_ms"].append(diagnostics.get("interpolated_exposure_ms"))
        data["target_gain"].append(diagnostics.get("target_gain"))
        data["interpolated_gain"].append(diagnostics.get("interpolated_gain"))
        data["transition_position"].append(diagnostics.get("transition_position"))
        data["sun_elevation"].append(diagnostics.get("sun_elevation"))

        # Brightness analysis data
        brightness = diagnostics.get("brightness", {})
        data["brightness_mean"].append(brightness.get("mean_brightness"))
        data["brightness_median"].append(brightness.get("median_brightness"))
        data["brightness_std"].append(brightness.get("std_brightness"))
        data["brightness_p5"].append(brightness.get("percentile_5"))
        data["brightness_p95"].append(brightness.get("percentile_95"))
        data["underexposed_percent"].append(brightness.get("underexposed_percent"))
        data["overexposed_percent"].append(brightness.get("overexposed_percent"))

    print(f"✅ Analysis complete! Collected {len(data['timestamps'])} data points.\n")

    # Check if diagnostic data is available
    has_diagnostics = any(m is not None for m in data["target_exposure_ms"])
    if has_diagnostics:
        print(
            f"📊 Diagnostic metadata found in {sum(1 for m in data['target_exposure_ms'] if m is not None)} images"
        )
    else:
        print("ℹ️  No diagnostic metadata found (older captures)")

    return data


def find_transition_zones(timestamps: List, modes: List) -> List[Tuple[datetime, datetime, str]]:
    """
    Find continuous time ranges for each mode (day, night, transition).

    Returns list of tuples: (start_time, end_time, mode)
    """
    if not timestamps or not modes:
        return []

    zones = []
    current_mode = modes[0] or "unknown"
    zone_start = timestamps[0]

    for i in range(1, len(timestamps)):
        mode = modes[i] or "unknown"
        if mode != current_mode:
            # End current zone, start new one
            zones.append((zone_start, timestamps[i - 1], current_mode))
            zone_start = timestamps[i]
            current_mode = mode

    # Add final zone
    zones.append((zone_start, timestamps[-1], current_mode))

    return zones


def create_graphs(data: Dict, output_dir: Path, config: dict):
    """
    Create and save analysis graphs focused on metadata.
    """
    if not data["timestamps"]:
        print("❌ No data to plot!")
        return

    print("📈 Creating graphs from metadata...")

    # Set up the style
    plt.style.use("seaborn-v0_8-darkgrid")

    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)

    # Common figure settings
    fig_width = 14
    fig_height = 8
    dpi = 150

    # 1. LUX - Dedicated graph (most important!)
    print("  Creating dedicated Lux graph...")
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    # Set dark background for prettier appearance
    fig.patch.set_facecolor("#1a1a1a")
    ax.set_facecolor("#2d2d2d")

    # Get thresholds for day/night coloring
    if "adaptive_timelapse" in config:
        thresholds = config["adaptive_timelapse"]["light_thresholds"]
        night_threshold = thresholds["night"]
        day_threshold = thresholds["day"]
    else:
        night_threshold = 10
        day_threshold = 100

    # Add colored background zones for day/night/twilight
    min_lux = min(data["lux"]) if data["lux"] else 1
    max_lux = max(data["lux"]) if data["lux"] else 100000

    # Night zone (dark blue)
    ax.axhspan(0.01, night_threshold, alpha=0.15, color="midnightblue", zorder=0)
    # Twilight zone (purple)
    ax.axhspan(night_threshold, day_threshold, alpha=0.15, color="mediumpurple", zorder=0)
    # Day zone (light yellow)
    ax.axhspan(day_threshold, max_lux * 10, alpha=0.10, color="gold", zorder=0)

    # Plot lux line with gradient-like effect using fill
    ax.plot(
        data["timestamps"],
        data["lux"],
        color="#ffaa00",
        linewidth=3,
        label="Measured Light",
        zorder=5,
        alpha=0.9,
    )

    # Add fill under the curve for prettier effect
    ax.fill_between(data["timestamps"], data["lux"], alpha=0.3, color="orange", zorder=4)

    # Style axes and labels with light colors for dark background
    ax.set_xlabel("Time", fontsize=14, fontweight="bold", color="white")
    ax.set_ylabel("Light Level (lux)", fontsize=14, fontweight="bold", color="white")
    title_text = (
        "Light Levels Over Time\n"
        "(Lux = unit of illuminance, 1 lux = light from 1 candle at 1 meter)"
    )
    ax.set_title(title_text, fontsize=16, fontweight="bold", pad=20, color="white")
    ax.set_yscale("log")

    # Grid styling
    ax.grid(True, alpha=0.2, linestyle="--", which="both", color="gray")
    ax.tick_params(colors="white", which="both")

    # Add reference lines for common light levels
    reference_levels = [
        (100000, "Direct sunlight", "#ff4444"),
        (10000, "Full daylight", "#ffaa44"),
        (1000, "Overcast day", "#ffff88"),
        (400, "Sunrise/Sunset", "#ff88cc"),
        (100, "Very dark day", "#aa88ff"),
        (10, "Twilight", "#6688ff"),
        (1, "Deep twilight", "#4444ff"),
        (0.1, "Full moon", "#888888"),
    ]

    for lux_val, label, color in reference_levels:
        if min_lux * 0.5 < lux_val < max_lux * 2:
            ax.axhline(
                y=lux_val,
                color=color,
                linestyle=":",
                linewidth=1.5,
                alpha=0.5,
                zorder=3,
            )
            ax.text(
                ax.get_xlim()[1],
                lux_val,
                f"  {label} ({lux_val:g})",
                fontsize=9,
                va="center",
                alpha=0.8,
                color=color,
                fontweight="bold",
            )

    # Add day/night threshold lines
    ax.axhline(
        y=night_threshold,
        color="#4444ff",
        linestyle="--",
        linewidth=2.5,
        alpha=0.9,
        label=f"Night threshold ({night_threshold} lux)",
        zorder=6,
    )
    ax.axhline(
        y=day_threshold,
        color="#ffdd44",
        linestyle="--",
        linewidth=2.5,
        alpha=0.9,
        label=f"Day threshold ({day_threshold} lux)",
        zorder=6,
    )

    # Legend with dark theme
    legend = ax.legend(
        loc="best", fontsize=11, framealpha=0.85, facecolor="#3d3d3d", edgecolor="gray"
    )
    plt.setp(legend.get_texts(), color="white")

    # Format x-axis
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

    # Spine colors
    for spine in ax.spines.values():
        spine.set_edgecolor("gray")
        spine.set_linewidth(1.5)

    fig.tight_layout()
    output_path = output_dir / "lux_levels.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # 2. Exposure Time
    print("  Creating exposure time graph...")
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    ax.plot(
        data["timestamps"],
        data["exposure_time"],
        color="tab:purple",
        linewidth=2,
        marker="o",
        markersize=3,
    )
    ax.set_xlabel("Time", fontsize=12)
    ax.set_ylabel("Exposure Time (seconds)", fontsize=12)
    ax.set_title("Camera Exposure Time Over Time", fontsize=16, fontweight="bold")
    ax.grid(True, alpha=0.3)

    # Add max exposure line from config
    if "adaptive_timelapse" in config:
        max_exp = config["adaptive_timelapse"]["night_mode"]["max_exposure_time"]
        ax.axhline(
            y=max_exp,
            color="red",
            linestyle="--",
            linewidth=1.5,
            alpha=0.7,
            label=f"Max night exposure ({max_exp}s)",
        )
        ax.legend(loc="upper left", fontsize=10)

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

    fig.tight_layout()
    output_path = output_dir / "exposure_time.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # 3. Analogue Gain (ISO)
    print("  Creating analogue gain graph...")
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    ax.plot(
        data["timestamps"],
        data["analogue_gain"],
        color="tab:green",
        linewidth=2,
        marker="o",
        markersize=3,
    )
    ax.set_xlabel("Time", fontsize=12)
    ax.set_ylabel("Analogue Gain (ISO equivalent)", fontsize=12)
    ax.set_title("Camera Analogue Gain Over Time", fontsize=16, fontweight="bold")
    ax.grid(True, alpha=0.3)

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

    fig.tight_layout()
    output_path = output_dir / "analogue_gain.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # 4. Sensor Temperature
    print("  Creating temperature graph...")
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    ax.plot(
        data["timestamps"],
        data["sensor_temp"],
        color="tab:red",
        linewidth=2,
        marker="o",
        markersize=3,
    )
    ax.set_xlabel("Time", fontsize=12)
    ax.set_ylabel("Temperature (°C)", fontsize=12)
    ax.set_title("Camera Sensor Temperature", fontsize=16, fontweight="bold")
    ax.grid(True, alpha=0.3)

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

    fig.tight_layout()
    output_path = output_dir / "temperature.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # 5. White Balance (Color Temperature and Gains)
    print("  Creating white balance graph...")
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(fig_width, fig_height * 1.2))

    # Color temperature
    ax1.plot(
        data["timestamps"],
        data["colour_temp"],
        color="tab:cyan",
        linewidth=2,
        marker="o",
        markersize=3,
    )
    ax1.set_ylabel("Color Temperature (K)", fontsize=11)
    ax1.set_title("White Balance - Color Temperature", fontsize=13, fontweight="bold")
    ax1.grid(True, alpha=0.3)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right")

    # Color gains (red and blue)
    ax2.plot(
        data["timestamps"],
        data["colour_gains_red"],
        color="red",
        linewidth=2,
        marker="o",
        markersize=3,
        label="Red Gain",
        alpha=0.7,
    )
    ax2.plot(
        data["timestamps"],
        data["colour_gains_blue"],
        color="blue",
        linewidth=2,
        marker="s",
        markersize=3,
        label="Blue Gain",
        alpha=0.7,
    )
    ax2.set_xlabel("Time", fontsize=11)
    ax2.set_ylabel("Color Gains", fontsize=11)
    ax2.set_title("White Balance - Color Gains (R/B)", fontsize=13, fontweight="bold")
    ax2.grid(True, alpha=0.3)
    ax2.legend(loc="upper left", fontsize=10)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha="right")

    fig.suptitle("White Balance Analysis", fontsize=16, fontweight="bold")
    fig.tight_layout()

    output_path = output_dir / "white_balance.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # 6. Overview Panel (4 key metrics)
    print("  Creating overview panel...")
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(fig_width * 1.2, fig_height * 1.2))

    # Lux (log scale)
    ax1.plot(data["timestamps"], data["lux"], color="tab:orange", linewidth=2)
    ax1.set_ylabel("Lux (log scale)", fontsize=10)
    ax1.set_title("Light Levels", fontsize=12, fontweight="bold")
    ax1.set_yscale("log")
    ax1.grid(True, alpha=0.3)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right")

    # Exposure
    ax2.plot(data["timestamps"], data["exposure_time"], color="tab:purple", linewidth=2)
    ax2.set_ylabel("Exposure (s)", fontsize=10)
    ax2.set_title("Exposure Time", fontsize=12, fontweight="bold")
    ax2.grid(True, alpha=0.3)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha="right")

    # Gain
    ax3.plot(data["timestamps"], data["analogue_gain"], color="tab:green", linewidth=2)
    ax3.set_ylabel("Gain", fontsize=10)
    ax3.set_xlabel("Time", fontsize=10)
    ax3.set_title("Analogue Gain (ISO)", fontsize=12, fontweight="bold")
    ax3.grid(True, alpha=0.3)
    ax3.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha="right")

    # Temperature
    ax4.plot(data["timestamps"], data["sensor_temp"], color="tab:red", linewidth=2)
    ax4.set_ylabel("Temperature (°C)", fontsize=10)
    ax4.set_xlabel("Time", fontsize=10)
    ax4.set_title("Sensor Temperature", fontsize=12, fontweight="bold")
    ax4.grid(True, alpha=0.3)
    ax4.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45, ha="right")

    fig.suptitle("Camera Metadata Overview", fontsize=16, fontweight="bold")
    fig.tight_layout()

    output_path = output_dir / "overview.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # === DIAGNOSTIC GRAPHS (if data available) ===
    has_diagnostics = any(m is not None for m in data["target_exposure_ms"])

    if has_diagnostics:
        print("\n📊 Creating diagnostic graphs...")

        # Filter out None values for plotting
        diag_timestamps = []
        diag_target_exp = []
        diag_actual_exp = []
        diag_target_gain = []
        diag_actual_gain = []
        diag_raw_lux = []
        diag_smoothed_lux = []

        for i, ts in enumerate(data["timestamps"]):
            if data["target_exposure_ms"][i] is not None:
                diag_timestamps.append(ts)
                diag_target_exp.append(data["target_exposure_ms"][i])
                diag_actual_exp.append(
                    data["interpolated_exposure_ms"][i] or data["target_exposure_ms"][i]
                )
                diag_target_gain.append(data["target_gain"][i] or 1.0)
                diag_actual_gain.append(
                    data["interpolated_gain"][i] or data["target_gain"][i] or 1.0
                )
                diag_raw_lux.append(data["raw_lux"][i] or data["lux"][i])
                diag_smoothed_lux.append(data["smoothed_lux"][i] or data["lux"][i])

        if diag_timestamps:
            # 7. Target vs Actual Exposure
            print("  Creating target vs actual exposure graph...")
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))

            ax.plot(
                diag_timestamps,
                diag_target_exp,
                color="tab:blue",
                linewidth=2,
                label="Target Exposure (calculated)",
                alpha=0.7,
            )
            ax.plot(
                diag_timestamps,
                diag_actual_exp,
                color="tab:orange",
                linewidth=2,
                label="Actual Exposure (interpolated)",
                alpha=0.9,
            )

            ax.set_xlabel("Time", fontsize=12)
            ax.set_ylabel("Exposure Time (ms)", fontsize=12)
            ax.set_title(
                "Exposure: Target vs Actual (Interpolated)\n"
                "Shows how smoothing affects exposure transitions",
                fontsize=14,
                fontweight="bold",
            )
            ax.grid(True, alpha=0.3)
            ax.legend(loc="upper left", fontsize=10)
            ax.set_yscale("log")

            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

            fig.tight_layout()
            output_path = output_dir / "exposure_target_vs_actual.png"
            plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
            plt.close()
            print(f"    ✅ Saved: {output_path}")

            # 8. Target vs Actual Gain
            print("  Creating target vs actual gain graph...")
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))

            ax.plot(
                diag_timestamps,
                diag_target_gain,
                color="tab:blue",
                linewidth=2,
                label="Target Gain (calculated)",
                alpha=0.7,
            )
            ax.plot(
                diag_timestamps,
                diag_actual_gain,
                color="tab:green",
                linewidth=2,
                label="Actual Gain (interpolated)",
                alpha=0.9,
            )

            ax.set_xlabel("Time", fontsize=12)
            ax.set_ylabel("Analogue Gain (ISO)", fontsize=12)
            ax.set_title(
                "Gain: Target vs Actual (Interpolated)\n"
                "Shows how smoothing affects ISO transitions",
                fontsize=14,
                fontweight="bold",
            )
            ax.grid(True, alpha=0.3)
            ax.legend(loc="upper left", fontsize=10)

            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

            fig.tight_layout()
            output_path = output_dir / "gain_target_vs_actual.png"
            plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
            plt.close()
            print(f"    ✅ Saved: {output_path}")

            # 9. Raw vs Smoothed Lux
            print("  Creating raw vs smoothed lux graph...")
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))

            ax.plot(
                diag_timestamps,
                diag_raw_lux,
                color="tab:gray",
                linewidth=1,
                label="Raw Lux (from test shot)",
                alpha=0.5,
            )
            ax.plot(
                diag_timestamps,
                diag_smoothed_lux,
                color="tab:orange",
                linewidth=2,
                label="Smoothed Lux (EMA)",
                alpha=0.9,
            )

            ax.set_xlabel("Time", fontsize=12)
            ax.set_ylabel("Lux (log scale)", fontsize=12)
            ax.set_title(
                "Lux: Raw vs Smoothed (EMA)\n" "Shows how lux smoothing reduces noise",
                fontsize=14,
                fontweight="bold",
            )
            ax.grid(True, alpha=0.3)
            ax.legend(loc="upper left", fontsize=10)
            ax.set_yscale("log")

            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            ax.xaxis.set_major_locator(mdates.HourLocator(interval=2))
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

            fig.tight_layout()
            output_path = output_dir / "lux_raw_vs_smoothed.png"
            plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
            plt.close()
            print(f"    ✅ Saved: {output_path}")

    # === BRIGHTNESS ANALYSIS GRAPHS (if data available) ===
    has_brightness = any(m is not None for m in data["brightness_mean"])

    if has_brightness:
        print("\n📊 Creating brightness analysis graphs...")

        # Filter out None values
        bright_timestamps = []
        bright_mean = []
        bright_p5 = []
        bright_p95 = []
        under_pct = []
        over_pct = []

        for i, ts in enumerate(data["timestamps"]):
            if data["brightness_mean"][i] is not None:
                bright_timestamps.append(ts)
                bright_mean.append(data["brightness_mean"][i])
                bright_p5.append(data["brightness_p5"][i] or 0)
                bright_p95.append(data["brightness_p95"][i] or 255)
                under_pct.append(data["underexposed_percent"][i] or 0)
                over_pct.append(data["overexposed_percent"][i] or 0)

        if bright_timestamps:
            # 10. Image Brightness Over Time
            print("  Creating image brightness graph...")
            fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(fig_width, fig_height * 1.2))

            # Top: Mean brightness with range
            ax1.fill_between(
                bright_timestamps,
                bright_p5,
                bright_p95,
                alpha=0.3,
                color="tab:blue",
                label="5th-95th percentile range",
            )
            ax1.plot(
                bright_timestamps,
                bright_mean,
                color="tab:blue",
                linewidth=2,
                label="Mean brightness",
            )

            # Add ideal range lines
            ax1.axhline(
                y=80,
                color="green",
                linestyle="--",
                linewidth=1,
                alpha=0.7,
                label="Ideal range (80-180)",
            )
            ax1.axhline(y=180, color="green", linestyle="--", linewidth=1, alpha=0.7)

            ax1.set_ylabel("Brightness (0-255)", fontsize=11)
            ax1.set_title("Image Brightness Over Time", fontsize=13, fontweight="bold")
            ax1.grid(True, alpha=0.3)
            ax1.legend(loc="upper right", fontsize=9)
            ax1.set_ylim(0, 255)
            ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right")

            # Bottom: Under/overexposed percentages
            ax2.fill_between(
                bright_timestamps,
                under_pct,
                alpha=0.5,
                color="tab:purple",
                label="Underexposed (<10)",
            )
            ax2.fill_between(
                bright_timestamps,
                over_pct,
                alpha=0.5,
                color="tab:red",
                label="Overexposed (>245)",
            )

            # Warning threshold
            ax2.axhline(
                y=5,
                color="orange",
                linestyle="--",
                linewidth=1,
                alpha=0.7,
                label="5% warning threshold",
            )

            ax2.set_xlabel("Time", fontsize=11)
            ax2.set_ylabel("Percentage (%)", fontsize=11)
            ax2.set_title("Under/Overexposed Pixels", fontsize=13, fontweight="bold")
            ax2.grid(True, alpha=0.3)
            ax2.legend(loc="upper right", fontsize=9)
            ax2.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha="right")

            fig.suptitle("Image Brightness Analysis", fontsize=16, fontweight="bold")
            fig.tight_layout()

            output_path = output_dir / "brightness_analysis.png"
            plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
            plt.close()
            print(f"    ✅ Saved: {output_path}")

            # 11. Diagnostic Overview Panel
            print("  Creating diagnostic overview panel...")
            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(
                2, 2, figsize=(fig_width * 1.2, fig_height * 1.2)
            )

            # Exposure target vs actual
            ax1.plot(
                diag_timestamps,
                diag_target_exp,
                color="tab:blue",
                linewidth=1.5,
                alpha=0.6,
                label="Target",
            )
            ax1.plot(
                diag_timestamps, diag_actual_exp, color="tab:orange", linewidth=2, label="Actual"
            )
            ax1.set_ylabel("Exposure (ms)", fontsize=10)
            ax1.set_title("Exposure: Target vs Actual", fontsize=12, fontweight="bold")
            ax1.set_yscale("log")
            ax1.grid(True, alpha=0.3)
            ax1.legend(loc="upper left", fontsize=8)
            ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right")

            # Gain target vs actual
            ax2.plot(
                diag_timestamps,
                diag_target_gain,
                color="tab:blue",
                linewidth=1.5,
                alpha=0.6,
                label="Target",
            )
            ax2.plot(
                diag_timestamps, diag_actual_gain, color="tab:green", linewidth=2, label="Actual"
            )
            ax2.set_ylabel("Gain (ISO)", fontsize=10)
            ax2.set_title("Gain: Target vs Actual", fontsize=12, fontweight="bold")
            ax2.grid(True, alpha=0.3)
            ax2.legend(loc="upper left", fontsize=8)
            ax2.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha="right")

            # Mean brightness
            ax3.fill_between(bright_timestamps, bright_p5, bright_p95, alpha=0.3, color="tab:blue")
            ax3.plot(bright_timestamps, bright_mean, color="tab:blue", linewidth=2)
            ax3.axhline(y=80, color="green", linestyle="--", linewidth=1, alpha=0.5)
            ax3.axhline(y=180, color="green", linestyle="--", linewidth=1, alpha=0.5)
            ax3.set_ylabel("Brightness", fontsize=10)
            ax3.set_xlabel("Time", fontsize=10)
            ax3.set_title("Image Brightness", fontsize=12, fontweight="bold")
            ax3.set_ylim(0, 255)
            ax3.grid(True, alpha=0.3)
            ax3.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha="right")

            # Under/overexposed
            ax4.fill_between(
                bright_timestamps, under_pct, alpha=0.5, color="tab:purple", label="Under"
            )
            ax4.fill_between(bright_timestamps, over_pct, alpha=0.5, color="tab:red", label="Over")
            ax4.axhline(y=5, color="orange", linestyle="--", linewidth=1, alpha=0.7)
            ax4.set_ylabel("Percent (%)", fontsize=10)
            ax4.set_xlabel("Time", fontsize=10)
            ax4.set_title("Clipped Pixels", fontsize=12, fontweight="bold")
            ax4.grid(True, alpha=0.3)
            ax4.legend(loc="upper right", fontsize=8)
            ax4.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45, ha="right")

            fig.suptitle("Diagnostic Overview", fontsize=16, fontweight="bold")
            fig.tight_layout()

            output_path = output_dir / "diagnostic_overview.png"
            plt.savefig(output_path, dpi=dpi, bbox_inches="tight")
            plt.close()
            print(f"    ✅ Saved: {output_path}")

    # === HOLY GRAIL TRANSITION ANALYSIS GRAPHS ===
    # These graphs help verify the smoothness of day/night transitions
    print("\n📊 Creating Holy Grail transition analysis graphs...")

    # Find transition zones from mode data
    transition_zones = find_transition_zones(data["timestamps"], data["mode"])
    has_transition_data = any(m == "transition" for m in data["mode"] if m)

    if has_transition_data:
        print(
            f"  Found {sum(1 for z in transition_zones if z[2] == 'transition')} transition periods"
        )

    # Mode colors for zone shading
    mode_colors = {
        "day": ("#ffff88", 0.15),  # Light yellow
        "night": ("#4444aa", 0.20),  # Dark blue
        "transition": ("#ff88ff", 0.25),  # Magenta/purple
        "unknown": ("#888888", 0.10),  # Gray
    }

    # Helper function to add zone shading to an axis
    def add_zone_shading(ax, zones, y_min, y_max):
        for start, end, mode in zones:
            color, alpha = mode_colors.get(mode, ("#888888", 0.1))
            ax.axvspan(start, end, alpha=alpha, color=color, zorder=0)

    # 12. HOLY GRAIL: Multi-Variable Comparison Plot (Lux vs Exposure vs Gain)
    print("  Creating Holy Grail multi-variable comparison plot...")
    fig, ax1 = plt.subplots(figsize=(fig_width, fig_height))

    # Set dark background
    fig.patch.set_facecolor("#1a1a1a")
    ax1.set_facecolor("#2d2d2d")

    # Add zone shading
    if transition_zones:
        y_min = min(data["lux"]) if data["lux"] else 0.1
        y_max = max(data["lux"]) if data["lux"] else 100000
        add_zone_shading(ax1, transition_zones, y_min, y_max)

    # Plot Lux (primary y-axis, log scale)
    color_lux = "#ffaa00"
    ax1.semilogy(
        data["timestamps"],
        data["lux"],
        color=color_lux,
        linewidth=2.5,
        label="Lux (light level)",
        zorder=5,
    )
    ax1.set_xlabel("Time", fontsize=12, color="white")
    ax1.set_ylabel("Lux (log scale)", fontsize=12, color=color_lux)
    ax1.tick_params(axis="y", labelcolor=color_lux)
    ax1.tick_params(axis="x", colors="white")

    # Create second y-axis for Exposure Time
    ax2 = ax1.twinx()
    color_exp = "#88ff88"
    # Convert exposure to milliseconds for better readability
    exposure_ms = [e * 1000 for e in data["exposure_time"]]
    ax2.semilogy(
        data["timestamps"],
        exposure_ms,
        color=color_exp,
        linewidth=2,
        linestyle="--",
        label="Exposure (ms)",
        zorder=4,
        alpha=0.8,
    )
    ax2.set_ylabel("Exposure Time (ms, log scale)", fontsize=12, color=color_exp)
    ax2.tick_params(axis="y", labelcolor=color_exp)

    # Create third y-axis for Analogue Gain
    ax3 = ax1.twinx()
    ax3.spines["right"].set_position(("axes", 1.12))  # Offset the third axis
    color_gain = "#ff8888"
    ax3.plot(
        data["timestamps"],
        data["analogue_gain"],
        color=color_gain,
        linewidth=2,
        linestyle=":",
        label="Gain (ISO)",
        zorder=3,
        alpha=0.8,
    )
    ax3.set_ylabel("Analogue Gain", fontsize=12, color=color_gain)
    ax3.tick_params(axis="y", labelcolor=color_gain)

    # Title and legend
    ax1.set_title(
        "Holy Grail Transition Analysis\n"
        "Verify: Exposure should follow Lux inversely with smooth S-curve (no spikes)",
        fontsize=14,
        fontweight="bold",
        color="white",
        pad=20,
    )

    # Combined legend
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    lines3, labels3 = ax3.get_legend_handles_labels()
    legend = ax1.legend(
        lines1 + lines2 + lines3,
        labels1 + labels2 + labels3,
        loc="upper right",
        fontsize=10,
        facecolor="#3d3d3d",
        edgecolor="gray",
    )
    plt.setp(legend.get_texts(), color="white")

    # Add mode legend
    if has_transition_data:
        from matplotlib.patches import Patch

        mode_patches = [
            Patch(facecolor="#ffff88", alpha=0.3, label="Day Mode"),
            Patch(facecolor="#ff88ff", alpha=0.4, label="Transition Mode"),
            Patch(facecolor="#4444aa", alpha=0.4, label="Night Mode"),
        ]
        legend2 = ax1.legend(
            handles=mode_patches,
            loc="upper left",
            fontsize=9,
            facecolor="#3d3d3d",
            edgecolor="gray",
        )
        plt.setp(legend2.get_texts(), color="white")
        ax1.add_artist(legend)  # Keep both legends

    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax1.xaxis.set_major_locator(mdates.HourLocator(interval=2))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right")
    ax1.grid(True, alpha=0.2, linestyle="--", color="gray")

    for spine in ax1.spines.values():
        spine.set_edgecolor("gray")

    fig.tight_layout()
    output_path = output_dir / "holy_grail_comparison.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # 13. HOLY GRAIL: AWB Gain Stability Plot
    print("  Creating Holy Grail AWB stability plot...")
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(fig_width, fig_height * 1.2))

    # Set dark background
    fig.patch.set_facecolor("#1a1a1a")
    ax1.set_facecolor("#2d2d2d")
    ax2.set_facecolor("#2d2d2d")

    # Add zone shading to both axes
    if transition_zones:
        red_min = min(data["colour_gains_red"]) if data["colour_gains_red"] else 1
        red_max = max(data["colour_gains_red"]) if data["colour_gains_red"] else 4
        blue_min = min(data["colour_gains_blue"]) if data["colour_gains_blue"] else 1
        blue_max = max(data["colour_gains_blue"]) if data["colour_gains_blue"] else 4
        add_zone_shading(ax1, transition_zones, red_min, red_max)
        add_zone_shading(ax2, transition_zones, blue_min, blue_max)

    # Plot Red Gain
    ax1.plot(
        data["timestamps"],
        data["colour_gains_red"],
        color="#ff4444",
        linewidth=2,
        marker="o",
        markersize=2,
        label="Red Gain",
    )
    ax1.set_ylabel("Red AWB Gain", fontsize=12, color="#ff4444")
    ax1.tick_params(axis="y", labelcolor="#ff4444")
    ax1.tick_params(axis="x", colors="white")
    ax1.set_title(
        "AWB Red Gain - Should be FLAT during Transition (purple zones)",
        fontsize=12,
        fontweight="bold",
        color="white",
    )
    ax1.grid(True, alpha=0.2, linestyle="--", color="gray")
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))

    # Calculate jitter (standard deviation) in transition zones
    transition_red_values = [
        data["colour_gains_red"][i]
        for i in range(len(data["timestamps"]))
        if data["mode"][i] == "transition"
    ]
    if transition_red_values:
        red_jitter = np.std(transition_red_values)
        ax1.text(
            0.02,
            0.98,
            f"Transition Jitter: {red_jitter:.4f}",
            transform=ax1.transAxes,
            fontsize=10,
            color="white",
            verticalalignment="top",
            bbox=dict(boxstyle="round", facecolor="#3d3d3d", alpha=0.8),
        )
        if red_jitter < 0.01:
            ax1.text(
                0.02,
                0.85,
                "✓ AWB LOCKED",
                transform=ax1.transAxes,
                fontsize=12,
                color="#88ff88",
                fontweight="bold",
                verticalalignment="top",
            )
        else:
            ax1.text(
                0.02,
                0.85,
                "⚠ AWB UNSTABLE",
                transform=ax1.transAxes,
                fontsize=12,
                color="#ff8888",
                fontweight="bold",
                verticalalignment="top",
            )

    # Plot Blue Gain
    ax2.plot(
        data["timestamps"],
        data["colour_gains_blue"],
        color="#4488ff",
        linewidth=2,
        marker="o",
        markersize=2,
        label="Blue Gain",
    )
    ax2.set_xlabel("Time", fontsize=12, color="white")
    ax2.set_ylabel("Blue AWB Gain", fontsize=12, color="#4488ff")
    ax2.tick_params(axis="y", labelcolor="#4488ff")
    ax2.tick_params(axis="x", colors="white")
    ax2.set_title(
        "AWB Blue Gain - Should be FLAT during Transition (purple zones)",
        fontsize=12,
        fontweight="bold",
        color="white",
    )
    ax2.grid(True, alpha=0.2, linestyle="--", color="gray")
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha="right")

    # Calculate jitter for blue
    transition_blue_values = [
        data["colour_gains_blue"][i]
        for i in range(len(data["timestamps"]))
        if data["mode"][i] == "transition"
    ]
    if transition_blue_values:
        blue_jitter = np.std(transition_blue_values)
        ax2.text(
            0.02,
            0.98,
            f"Transition Jitter: {blue_jitter:.4f}",
            transform=ax2.transAxes,
            fontsize=10,
            color="white",
            verticalalignment="top",
            bbox=dict(boxstyle="round", facecolor="#3d3d3d", alpha=0.8),
        )
        if blue_jitter < 0.01:
            ax2.text(
                0.02,
                0.85,
                "✓ AWB LOCKED",
                transform=ax2.transAxes,
                fontsize=12,
                color="#88ff88",
                fontweight="bold",
                verticalalignment="top",
            )
        else:
            ax2.text(
                0.02,
                0.85,
                "⚠ AWB UNSTABLE",
                transform=ax2.transAxes,
                fontsize=12,
                color="#ff8888",
                fontweight="bold",
                verticalalignment="top",
            )

    for spine in ax1.spines.values():
        spine.set_edgecolor("gray")
    for spine in ax2.spines.values():
        spine.set_edgecolor("gray")

    fig.suptitle(
        "Holy Grail AWB Stability Check\n"
        "Verify: Gains should be perfectly flat (no jitter) in purple transition zones",
        fontsize=14,
        fontweight="bold",
        color="white",
    )
    fig.tight_layout()

    output_path = output_dir / "holy_grail_awb_stability.png"
    plt.savefig(output_path, dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    print(f"    ✅ Saved: {output_path}")

    # 14. HOLY GRAIL: Transition Position / Factor Plot
    # Shows the transition_position (0.0 = night threshold, 1.0 = day threshold)
    transition_positions = [
        (data["timestamps"][i], data["transition_position"][i])
        for i in range(len(data["timestamps"]))
        if data["transition_position"][i] is not None
    ]

    if transition_positions:
        print("  Creating transition position plot...")
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))

        fig.patch.set_facecolor("#1a1a1a")
        ax.set_facecolor("#2d2d2d")

        # Add zone shading
        if transition_zones:
            add_zone_shading(ax, transition_zones, 0, 1)

        trans_times = [t[0] for t in transition_positions]
        trans_values = [t[1] for t in transition_positions]

        # Plot transition position
        ax.plot(
            trans_times,
            trans_values,
            color="#ff88ff",
            linewidth=3,
            marker="o",
            markersize=4,
            label="Transition Position",
        )
        ax.fill_between(trans_times, trans_values, alpha=0.3, color="#ff88ff")

        # Add reference lines
        ax.axhline(
            y=0.0,
            color="#4444ff",
            linestyle="--",
            linewidth=1.5,
            alpha=0.7,
            label="Night threshold (0.0)",
        )
        ax.axhline(
            y=1.0,
            color="#ffdd44",
            linestyle="--",
            linewidth=1.5,
            alpha=0.7,
            label="Day threshold (1.0)",
        )
        ax.axhline(
            y=0.5, color="#888888", linestyle=":", linewidth=1, alpha=0.5, label="Midpoint (0.5)"
        )

        ax.set_xlabel("Time", fontsize=12, color="white")
        ax.set_ylabel("Transition Position (0=Night, 1=Day)", fontsize=12, color="white")
        ax.set_title(
            "Holy Grail Transition Factor\n"
            "Shows progress through twilight zone - should be smooth S-curve, not Z-shape",
            fontsize=14,
            fontweight="bold",
            color="white",
            pad=20,
        )
        ax.set_ylim(-0.1, 1.1)
        ax.tick_params(colors="white")
        ax.grid(True, alpha=0.2, linestyle="--", color="gray")

        legend = ax.legend(loc="upper right", fontsize=10, facecolor="#3d3d3d", edgecolor="gray")
        plt.setp(legend.get_texts(), color="white")

        ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        ax.xaxis.set_major_locator(mdates.HourLocator(interval=1))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

        for spine in ax.spines.values():
            spine.set_edgecolor("gray")

        fig.tight_layout()
        output_path = output_dir / "holy_grail_transition_factor.png"
        plt.savefig(output_path, dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close()
        print(f"    ✅ Saved: {output_path}")
    else:
        print("  ℹ️  No transition position data available (diagnostics may be disabled)")

    # 15. POLAR: Sun Elevation + Lux Comparison
    # Shows sun position alongside lux for polar region analysis
    sun_data = [
        (data["timestamps"][i], data["sun_elevation"][i])
        for i in range(len(data["timestamps"]))
        if data["sun_elevation"][i] is not None
    ]

    if sun_data:
        print("  Creating sun elevation plot...")
        fig, ax1 = plt.subplots(figsize=(fig_width, fig_height))

        fig.patch.set_facecolor("#1a1a1a")
        ax1.set_facecolor("#2d2d2d")

        # Add zone shading
        if transition_zones:
            add_zone_shading(ax1, transition_zones, -20, 50)

        sun_times = [t[0] for t in sun_data]
        sun_values = [t[1] for t in sun_data]

        # Plot Sun Elevation (primary axis)
        color_sun = "#ffdd44"
        ax1.plot(
            sun_times,
            sun_values,
            color=color_sun,
            linewidth=2.5,
            label="Sun Elevation (°)",
            zorder=5,
        )
        ax1.fill_between(
            sun_times,
            sun_values,
            alpha=0.2,
            color=color_sun,
            zorder=4,
        )

        # Add reference lines for sun position
        ax1.axhline(
            y=0,
            color="#ff8844",
            linestyle="-",
            linewidth=2,
            alpha=0.7,
            label="Horizon (0°)",
        )
        ax1.axhline(
            y=-6,
            color="#8888ff",
            linestyle="--",
            linewidth=2,
            alpha=0.7,
            label="Civil Twilight (-6°)",
        )
        ax1.axhline(
            y=-12,
            color="#4444ff",
            linestyle=":",
            linewidth=1.5,
            alpha=0.5,
            label="Nautical Twilight (-12°)",
        )

        ax1.set_xlabel("Time", fontsize=12, color="white")
        ax1.set_ylabel("Sun Elevation (degrees)", fontsize=12, color=color_sun)
        ax1.tick_params(axis="y", labelcolor=color_sun)
        ax1.tick_params(axis="x", colors="white")

        # Create second axis for Lux
        ax2 = ax1.twinx()
        color_lux = "#88ff88"

        # Match lux data timestamps to sun data
        lux_for_sun = []
        for t, _ in sun_data:
            # Find closest timestamp in lux data
            idx = min(
                range(len(data["timestamps"])),
                key=lambda i: abs((data["timestamps"][i] - t).total_seconds()),
            )
            lux_for_sun.append(data["lux"][idx])

        ax2.semilogy(
            sun_times,
            lux_for_sun,
            color=color_lux,
            linewidth=2,
            linestyle="--",
            label="Lux (light)",
            alpha=0.8,
            zorder=3,
        )
        ax2.set_ylabel("Lux (log scale)", fontsize=12, color=color_lux)
        ax2.tick_params(axis="y", labelcolor=color_lux)

        ax1.set_title(
            "Polar Awareness: Sun Elevation vs Light Level\n"
            "When sun > -6° (civil twilight), Day Mode is forced for twilight colors",
            fontsize=14,
            fontweight="bold",
            color="white",
            pad=20,
        )

        # Combined legend
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        legend = ax1.legend(
            lines1 + lines2,
            labels1 + labels2,
            loc="upper right",
            fontsize=10,
            facecolor="#3d3d3d",
            edgecolor="gray",
        )
        plt.setp(legend.get_texts(), color="white")

        ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        ax1.xaxis.set_major_locator(mdates.HourLocator(interval=2))
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right")
        ax1.grid(True, alpha=0.2, linestyle="--", color="gray")

        for spine in ax1.spines.values():
            spine.set_edgecolor("gray")

        fig.tight_layout()
        output_path = output_dir / "polar_sun_elevation.png"
        plt.savefig(output_path, dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
        plt.close()
        print(f"    ✅ Saved: {output_path}")
    else:
        print("  ℹ️  No sun elevation data available (location may not be configured)")

    print("\n✅ All graphs created successfully!")


def print_statistics(data: Dict, hours: int):
    """Print statistical summary of the data."""
    if not data["timestamps"]:
        return

    print("\n" + "=" * 60)
    print(f"📊 STATISTICAL SUMMARY (Last {hours} hours)")
    print("=" * 60)

    print(f"\n🕒 Time Range:")
    print(f"  From: {min(data['timestamps']).strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  To:   {max(data['timestamps']).strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Total images: {len(data['timestamps'])}")

    print(f"\n💡 Light Levels (Lux):")
    lux_values = [l for l in data["lux"] if l > 0]  # Filter out zeros
    if lux_values:
        print(f"  Min:     {min(lux_values):.2f} lux")
        print(f"  Max:     {max(lux_values):.2f} lux")
        print(f"  Average: {np.mean(lux_values):.2f} lux")
        print(f"  Median:  {np.median(lux_values):.2f} lux")

    print(f"\n⏱️  Exposure Time (seconds):")
    print(f"  Min:     {min(data['exposure_time']):.4f}s")
    print(f"  Max:     {max(data['exposure_time']):.4f}s")
    print(f"  Average: {np.mean(data['exposure_time']):.4f}s")
    print(f"  Median:  {np.median(data['exposure_time']):.4f}s")

    print(f"\n📸 Analogue Gain (ISO):")
    print(f"  Min:     {min(data['analogue_gain']):.2f}")
    print(f"  Max:     {max(data['analogue_gain']):.2f}")
    print(f"  Average: {np.mean(data['analogue_gain']):.2f}")
    print(f"  Median:  {np.median(data['analogue_gain']):.2f}")

    print(f"\n🌡️  Sensor Temperature (°C):")
    print(f"  Min:     {min(data['sensor_temp']):.1f}°C")
    print(f"  Max:     {max(data['sensor_temp']):.1f}°C")
    print(f"  Average: {np.mean(data['sensor_temp']):.1f}°C")

    print(f"\n🎨 Color Temperature (K):")
    print(f"  Min:     {min(data['colour_temp'])}K")
    print(f"  Max:     {max(data['colour_temp'])}K")
    print(f"  Average: {np.mean(data['colour_temp']):.0f}K")

    print("\n" + "=" * 60)


def export_to_excel(
    data: Dict,
    output_path: Path,
    hours: int,
    config: dict,
    image_pairs: List[Tuple[Path, Path]],
):
    """Export analysis data to Excel file with multiple sheets."""
    print(f"\n📊 Creating Excel file: {output_path}")

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # === Sheet 1: Raw Data ===
    ws_raw = wb.create_sheet("Raw Data")

    # Headers
    headers = [
        "Timestamp",
        "Date",
        "Time",
        "Lux",
        "Exposure (s)",
        "Analogue Gain",
        "Sensor Temp (°C)",
        "Color Temp (K)",
        "Color Gain Red",
        "Color Gain Blue",
        "Digital Gain",
        "Mode",
        "Target Exp (ms)",
        "Actual Exp (ms)",
        "Target Gain",
        "Actual Gain",
        "Mean Brightness",
        "Under %",
        "Over %",
        "Image File",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for i, ts in enumerate(data["timestamps"], 2):
        idx = i - 2
        ws_raw.cell(row=i, column=1, value=ts.strftime("%Y-%m-%d %H:%M:%S"))
        ws_raw.cell(row=i, column=2, value=ts.strftime("%Y-%m-%d"))
        ws_raw.cell(row=i, column=3, value=ts.strftime("%H:%M:%S"))
        ws_raw.cell(row=i, column=4, value=round(data["lux"][idx], 2))
        ws_raw.cell(row=i, column=5, value=round(data["exposure_time"][idx], 4))
        ws_raw.cell(row=i, column=6, value=round(data["analogue_gain"][idx], 2))
        ws_raw.cell(row=i, column=7, value=round(data["sensor_temp"][idx], 1))
        ws_raw.cell(row=i, column=8, value=data["colour_temp"][idx])
        ws_raw.cell(row=i, column=9, value=round(data["colour_gains_red"][idx], 3))
        ws_raw.cell(row=i, column=10, value=round(data["colour_gains_blue"][idx], 3))
        ws_raw.cell(row=i, column=11, value=round(data["digital_gain"][idx], 3))
        # Diagnostic data
        ws_raw.cell(row=i, column=12, value=data["mode"][idx] or "")
        ws_raw.cell(
            row=i,
            column=13,
            value=(
                round(data["target_exposure_ms"][idx], 2) if data["target_exposure_ms"][idx] else ""
            ),
        )
        ws_raw.cell(
            row=i,
            column=14,
            value=(
                round(data["interpolated_exposure_ms"][idx], 2)
                if data["interpolated_exposure_ms"][idx]
                else ""
            ),
        )
        ws_raw.cell(
            row=i,
            column=15,
            value=round(data["target_gain"][idx], 2) if data["target_gain"][idx] else "",
        )
        ws_raw.cell(
            row=i,
            column=16,
            value=(
                round(data["interpolated_gain"][idx], 2) if data["interpolated_gain"][idx] else ""
            ),
        )
        ws_raw.cell(
            row=i,
            column=17,
            value=round(data["brightness_mean"][idx], 1) if data["brightness_mean"][idx] else "",
        )
        ws_raw.cell(
            row=i,
            column=18,
            value=(
                round(data["underexposed_percent"][idx], 2)
                if data["underexposed_percent"][idx]
                else ""
            ),
        )
        ws_raw.cell(
            row=i,
            column=19,
            value=(
                round(data["overexposed_percent"][idx], 2)
                if data["overexposed_percent"][idx]
                else ""
            ),
        )
        ws_raw.cell(row=i, column=20, value=data["filenames"][idx])

    # Auto-size columns
    for column in ws_raw.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    # === Sheet 2: Statistics ===
    ws_stats = wb.create_sheet("Statistics")

    ws_stats.cell(row=1, column=1, value="STATISTICAL SUMMARY").font = Font(bold=True, size=14)
    ws_stats.cell(row=2, column=1, value=f"Last {hours} hours")

    row = 4

    # Time range
    ws_stats.cell(row=row, column=1, value="TIME RANGE").font = Font(bold=True)
    row += 1
    ws_stats.cell(row=row, column=1, value="From:")
    ws_stats.cell(row=row, column=2, value=min(data["timestamps"]).strftime("%Y-%m-%d %H:%M:%S"))
    row += 1
    ws_stats.cell(row=row, column=1, value="To:")
    ws_stats.cell(row=row, column=2, value=max(data["timestamps"]).strftime("%Y-%m-%d %H:%M:%S"))
    row += 1
    ws_stats.cell(row=row, column=1, value="Total Images:")
    ws_stats.cell(row=row, column=2, value=len(data["timestamps"]))
    row += 2

    # Lux statistics
    lux_values = [l for l in data["lux"] if l > 0]
    if lux_values:
        ws_stats.cell(row=row, column=1, value="LIGHT LEVELS (Lux)").font = Font(bold=True)
        row += 1
        ws_stats.cell(row=row, column=1, value="Min:").fill = PatternFill(
            start_color="E0E0E0", fill_type="solid"
        )
        ws_stats.cell(row=row, column=2, value=round(min(lux_values), 2))
        row += 1
        ws_stats.cell(row=row, column=1, value="Max:").fill = PatternFill(
            start_color="E0E0E0", fill_type="solid"
        )
        ws_stats.cell(row=row, column=2, value=round(max(lux_values), 2))
        row += 1
        ws_stats.cell(row=row, column=1, value="Average:").fill = PatternFill(
            start_color="E0E0E0", fill_type="solid"
        )
        ws_stats.cell(row=row, column=2, value=round(np.mean(lux_values), 2))
        row += 1
        ws_stats.cell(row=row, column=1, value="Median:").fill = PatternFill(
            start_color="E0E0E0", fill_type="solid"
        )
        ws_stats.cell(row=row, column=2, value=round(np.median(lux_values), 2))
        row += 2

    # Exposure statistics
    ws_stats.cell(row=row, column=1, value="EXPOSURE TIME (seconds)").font = Font(bold=True)
    row += 1
    ws_stats.cell(row=row, column=1, value="Min:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(min(data["exposure_time"]), 4))
    row += 1
    ws_stats.cell(row=row, column=1, value="Max:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(max(data["exposure_time"]), 4))
    row += 1
    ws_stats.cell(row=row, column=1, value="Average:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(np.mean(data["exposure_time"]), 4))
    row += 1
    ws_stats.cell(row=row, column=1, value="Median:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(np.median(data["exposure_time"]), 4))
    row += 2

    # Gain statistics
    ws_stats.cell(row=row, column=1, value="ANALOGUE GAIN (ISO)").font = Font(bold=True)
    row += 1
    ws_stats.cell(row=row, column=1, value="Min:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(min(data["analogue_gain"]), 2))
    row += 1
    ws_stats.cell(row=row, column=1, value="Max:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(max(data["analogue_gain"]), 2))
    row += 1
    ws_stats.cell(row=row, column=1, value="Average:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(np.mean(data["analogue_gain"]), 2))
    row += 1
    ws_stats.cell(row=row, column=1, value="Median:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(np.median(data["analogue_gain"]), 2))
    row += 2

    # Temperature statistics
    ws_stats.cell(row=row, column=1, value="SENSOR TEMPERATURE (°C)").font = Font(bold=True)
    row += 1
    ws_stats.cell(row=row, column=1, value="Min:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(min(data["sensor_temp"]), 1))
    row += 1
    ws_stats.cell(row=row, column=1, value="Max:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(max(data["sensor_temp"]), 1))
    row += 1
    ws_stats.cell(row=row, column=1, value="Average:").fill = PatternFill(
        start_color="E0E0E0", fill_type="solid"
    )
    ws_stats.cell(row=row, column=2, value=round(np.mean(data["sensor_temp"]), 1))

    # Auto-size columns
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 25

    # === Sheet 3: Hourly Averages ===
    ws_hourly = wb.create_sheet("Hourly Averages")

    # Group data by hour
    hourly_data = {}
    for i, ts in enumerate(data["timestamps"]):
        hour_key = ts.replace(minute=0, second=0, microsecond=0)
        if hour_key not in hourly_data:
            hourly_data[hour_key] = {
                "lux": [],
                "exposure_time": [],
                "analogue_gain": [],
                "sensor_temp": [],
                "colour_temp": [],
            }
        hourly_data[hour_key]["lux"].append(data["lux"][i])
        hourly_data[hour_key]["exposure_time"].append(data["exposure_time"][i])
        hourly_data[hour_key]["analogue_gain"].append(data["analogue_gain"][i])
        hourly_data[hour_key]["sensor_temp"].append(data["sensor_temp"][i])
        hourly_data[hour_key]["colour_temp"].append(data["colour_temp"][i])

    # Headers
    hourly_headers = [
        "Hour",
        "Avg Lux",
        "Avg Exposure (s)",
        "Avg Gain",
        "Avg Temp (°C)",
        "Avg Color Temp (K)",
        "Image Count",
    ]

    for col, header in enumerate(hourly_headers, 1):
        cell = ws_hourly.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Data rows
    for row_idx, (hour, values) in enumerate(sorted(hourly_data.items()), 2):
        ws_hourly.cell(row=row_idx, column=1, value=hour.strftime("%Y-%m-%d %H:00"))
        ws_hourly.cell(
            row=row_idx,
            column=2,
            value=round(
                (
                    np.mean([l for l in values["lux"] if l > 0])
                    if any(l > 0 for l in values["lux"])
                    else 0
                ),
                2,
            ),
        )
        ws_hourly.cell(row=row_idx, column=3, value=round(np.mean(values["exposure_time"]), 4))
        ws_hourly.cell(row=row_idx, column=4, value=round(np.mean(values["analogue_gain"]), 2))
        ws_hourly.cell(row=row_idx, column=5, value=round(np.mean(values["sensor_temp"]), 1))
        ws_hourly.cell(row=row_idx, column=6, value=round(np.mean(values["colour_temp"]), 0))
        ws_hourly.cell(row=row_idx, column=7, value=len(values["lux"]))

    # Auto-size columns
    for column in ws_hourly.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_hourly.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel file saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Analyze timelapse images and generate graphs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Analyze last 24 hours (default)
  python3 src/analyze_timelapse.py

  # Analyze last 48 hours
  python3 src/analyze_timelapse.py --hours 48

  # Use custom config
  python3 src/analyze_timelapse.py -c config/custom.yml
        """,
    )

    parser.add_argument(
        "-c",
        "--config",
        default="config/config.yml",
        help="Path to config file (default: config/config.yml)",
    )

    parser.add_argument(
        "--hours", type=int, default=24, help="Number of hours to analyze (default: 24)"
    )

    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("  🎥 RASPILAPSE ANALYSIS TOOL 🎥")
    print("=" * 60)

    # Load configuration
    try:
        config = load_config(args.config)
    except Exception as e:
        print(f"❌ Error loading config: {e}")
        sys.exit(1)

    # Get output directory from config
    output_dir = Path(config["output"]["directory"])

    # Get graphs directory from config (or use default)
    graphs_dir = Path(config.get("graphs", {}).get("directory", "graphs"))

    print(f"\n⚙️  Configuration:")
    print(f"  Config file:   {args.config}")
    print(f"  Image dir:     {output_dir}")
    print(f"  Graphs dir:    {graphs_dir}")
    print(f"  Time window:   Last {args.hours} hours")

    # Find recent images
    print(f"\n🔍 Searching for images in {output_dir}...")
    image_pairs = find_recent_images(output_dir, args.hours)

    if not image_pairs:
        print(f"❌ No images found in the last {args.hours} hours!")
        print(f"   Check that images exist in: {output_dir}")
        sys.exit(1)

    print(f"✅ Found {len(image_pairs)} images with metadata")

    # Analyze images
    data = analyze_images(image_pairs, args.hours)

    # Print statistics
    print_statistics(data, args.hours)

    # Create graphs
    create_graphs(data, graphs_dir, config)

    # Export to Excel
    excel_filename = f"timelapse_analysis_{args.hours}h.xlsx"
    excel_path = graphs_dir / excel_filename
    export_to_excel(data, excel_path, args.hours, config, image_pairs)

    print(f"\n✅ Analysis complete!")
    print(f"\n📊 Graphs saved to: {graphs_dir}")
    print(f"   View your graphs:")
    for graph in sorted(graphs_dir.glob("*.png")):
        print(f"     - {graph.name}")
    print(f"\n📁 Excel file saved: {excel_path.name}")
    print(f"   Open with: libreoffice {excel_path}")
    print()


if __name__ == "__main__":
    main()
